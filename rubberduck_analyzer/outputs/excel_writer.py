"""Excel workbook generation for master analysis spreadsheet."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, row: int = 1):
    """Style the header row of a worksheet."""
    for cell in ws[row]:
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, min_width: int = 12, max_width: int = 50):
    """Auto-adjust column widths based on content."""
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def _write_tester_summary(wb: Workbook, sessions: list[dict]):
    """Write the Tester Summary sheet — one row per tester."""
    ws = wb.active
    ws.title = "Tester Summary"

    headers = [
        "Tester", "Date", "Milestone", "IDE", "Codebase", "Language", "Type",
        "Industry", "Size", "Setup Method", "Setup Duration (min)",
        "Prompt Independence", "MCP Usage", "Review Depth", "Trust Score",
        "M3 Candidacy", "Engagement", "Feature Requests", "Complaints",
    ]
    ws.append(headers)
    _style_header(ws)

    for session in sessions:
        tester = session.get("tester", {})
        codebase = tester.get("codebase", {})
        obs = session.get("observations", {})
        inst = obs.get("installation", {})
        prompt = obs.get("prompting", {})
        review = obs.get("output_review", {})
        trust = obs.get("trust", {})
        feedback = obs.get("product_feedback", {})
        m3 = session.get("m3_candidacy", {})

        ws.append([
            tester.get("name", ""),
            tester.get("date", ""),
            session.get("milestone", "M1"),
            tester.get("ide", ""),
            codebase.get("name", ""),
            codebase.get("language", ""),
            codebase.get("type", ""),
            codebase.get("industry", ""),
            codebase.get("size", ""),
            inst.get("setup_method", ""),
            inst.get("setup_duration_minutes", ""),
            prompt.get("prompt_independence", ""),
            prompt.get("mcp_tool_usage", ""),
            review.get("review_depth", ""),
            trust.get("trust_score", ""),
            m3.get("rating", ""),
            m3.get("engagement_level", ""),
            len(feedback.get("feature_requests", [])),
            len(feedback.get("complaints", [])),
        ])

    _auto_width(ws)


def _write_observation_detail(wb: Workbook, sessions: list[dict]):
    """Write the Observation Detail sheet with verbatim quotes."""
    ws = wb.create_sheet("Observation Detail")

    headers = ["Tester", "Area", "Key", "Value", "Verbatim Quotes"]
    ws.append(headers)
    _style_header(ws)

    for session in sessions:
        name = session.get("tester", {}).get("name", "unknown")
        obs = session.get("observations", {})
        for area_name, area_data in obs.items():
            if not isinstance(area_data, dict):
                continue
            quotes = area_data.get("verbatim_quotes", [])
            quotes_str = "\n".join(f'"{q}"' for q in quotes) if quotes else ""
            for key, value in area_data.items():
                if key == "verbatim_quotes":
                    continue
                val_str = json.dumps(value) if isinstance(value, (list, dict)) else str(value)
                ws.append([name, area_name, key, val_str, quotes_str])
                # Only show quotes on first row per area
                quotes_str = ""

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT

    _auto_width(ws)


def _write_product_feedback(wb: Workbook, sessions: list[dict]):
    """Write the Product Feedback sheet — deduplicated feature requests and complaints."""
    ws = wb.create_sheet("Product Feedback")

    headers = ["Type", "Description", "Priority/Severity", "Category", "Tester", "Competitor"]
    ws.append(headers)
    _style_header(ws)

    for session in sessions:
        name = session.get("tester", {}).get("name", "unknown")
        feedback = session.get("observations", {}).get("product_feedback", {})

        for req in feedback.get("feature_requests", []):
            ws.append([
                "Feature Request",
                req.get("description", ""),
                req.get("priority", ""),
                req.get("category", ""),
                name,
                "",
            ])

        for complaint in feedback.get("complaints", []):
            ws.append([
                "Complaint",
                complaint.get("description", ""),
                complaint.get("severity", ""),
                "",
                name,
                "",
            ])

        for comp in feedback.get("comparisons_to_competitors", []):
            ws.append([
                "Competitor Comparison",
                comp.get("feature", ""),
                comp.get("verdict", ""),
                "",
                name,
                comp.get("competitor", ""),
            ])

    _auto_width(ws)


def _write_evidence_tracker(wb: Workbook, sessions: list[dict]):
    """Write the Evidence Tracker sheet — M3 evidence classified and rated."""
    ws = wb.create_sheet("Evidence Tracker")

    headers = [
        "Tester", "Evidence Type", "Quality (1-5)", "RD Advantage",
        "Metrics Touched", "Marketing Usable", "Investor Usable",
        "RD Findings", "Normal Tools Missed",
    ]
    ws.append(headers)
    _style_header(ws)

    for session in sessions:
        if session.get("milestone") != "M3":
            continue
        name = session.get("tester_name", "unknown")
        ev = session.get("evidence", {})
        ws.append([
            name,
            ev.get("evidence_type", ""),
            ev.get("evidence_quality", ""),
            ev.get("rubberduck_advantage_demonstrated", ""),
            ", ".join(ev.get("metrics_touched", [])),
            ev.get("usable_for_marketing", ""),
            ev.get("usable_for_investor_deck", ""),
            "\n".join(ev.get("specific_findings_rubberduck_surfaced", [])),
            "\n".join(ev.get("specific_findings_normal_tools_missed", [])),
        ])

    _auto_width(ws)


def _write_patterns(wb: Workbook, patterns: dict):
    """Write the Patterns sheet — cross-tester patterns."""
    ws = wb.create_sheet("Patterns")

    headers = ["Category", "Pattern", "Value", "Supporting Data"]
    ws.append(headers)
    _style_header(ws)

    for category, cat_patterns in patterns.items():
        if isinstance(cat_patterns, dict):
            for key, value in cat_patterns.items():
                val_str = json.dumps(value) if isinstance(value, (list, dict)) else str(value)
                ws.append([category, key, val_str, ""])
        elif isinstance(cat_patterns, list):
            for item in cat_patterns:
                if isinstance(item, dict):
                    ws.append([category, item.get("pattern", ""), item.get("value", ""), item.get("data", "")])
                else:
                    ws.append([category, str(item), "", ""])

    _auto_width(ws)


def _write_action_items(wb: Workbook, action_items: list[dict]):
    """Write the Action Items sheet."""
    ws = wb.create_sheet("Action Items")

    headers = ["Priority", "Category", "Description", "Source", "Frequency"]
    ws.append(headers)
    _style_header(ws)

    for item in action_items:
        ws.append([
            item.get("priority", ""),
            item.get("category", ""),
            item.get("description", ""),
            item.get("source", ""),
            item.get("frequency", ""),
        ])

    _auto_width(ws)


def generate_workbook(
    sessions: list[dict],
    patterns: dict | None = None,
    action_items: list[dict] | None = None,
    output_path: str | Path = "data/reports/analysis.xlsx",
) -> Path:
    """Generate the master Excel workbook with all analysis sheets.

    Args:
        sessions: List of session JSON dicts (M1, M2, M3).
        patterns: Cross-tester pattern dict from synthesizer.
        action_items: Prioritized action items list.
        output_path: Where to write the workbook.

    Returns:
        Path to the generated workbook.
    """
    wb = Workbook()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    _write_tester_summary(wb, sessions)
    _write_observation_detail(wb, sessions)
    _write_product_feedback(wb, sessions)
    _write_evidence_tracker(wb, sessions)

    if patterns:
        _write_patterns(wb, patterns)
    if action_items:
        _write_action_items(wb, action_items)

    wb.save(str(output_path))
    return output_path
